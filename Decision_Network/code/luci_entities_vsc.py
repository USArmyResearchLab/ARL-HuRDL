#!/usr/bin/python

import random
import numpy
import luci, decision_net, randomize_world, UtilityCalculation
import xml.etree.ElementTree as ET
from xml.dom import minidom
from itertools import permutations
from itertools import combinations
from pysat.examples.hitman import Hitman
from openpyxl import Workbook
from openpyxl import load_workbook

import csv

# knowledge of entity features (list of dictionaries (string: list of strings)); this updates the decision net nodes
propKnowledge = {}
questionCount = {"Query:Location": 0, "Query:ID": 0, "Query:Color": 0, 
"Query:General": 0, "Query:Pattern": 0, "Query:Symbol": 0, 
"Confirm:Spatial": 0, "Confirm:Color": 0, "Confirm:Landmark": 0}
unscannable_props = ["Name"]#,"ContainedIn","Coord.X","Coord.Y"]

propToQ = {"Location": ["Query:Location", "Confirm:Spatial", "Confirm:Landmark"], 
    "ObjectType": ["Query:ID", "Query:General"], "Color": ["Query:Color", "Confirm:Color"], 
    "Pattern": ["Query:Pattern"], "Symbol": ["Query:Symbol"], "Shape": ["Query:Shape"],
    "Size": ["Query:Size"], "Texture": ["Query:Texture"]}

numInputUtts = 0 # tracks total # of input utts for mean calculation
qToProp = 0 # tracks q's to learn a prop

# flag for Random world 
random_world = False

# flag for complexity of random world ("high"/"low")
complexity = "high"

# set to "baseline" or "decision" (net) for question generation
condition = "decision" 

# set to "corpus" or "entropy" for utility assignment
utilities = "entropy"

class TestEntities(object):
    def __init__(self):
        self.entities = []

    def parseFromCSV(self,csvfilename):
        with open(csvfilename) as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                self.entities.append(row)

class Entropy(object):
    def __init__(self,WHQentropy,YNQentropy):
        self.WHQentropy = WHQentropy
        self.YNQentropy = YNQentropy


def computeUtilityTable(propList,questions,entropy):
    corpusUtilities = {"Query:Location": 32, "Confirm:Spatial": 10, "Confirm:Landmark": 4,
    "Query:ID": 16, "Query:General": 13, "Query:Color": 14, "Confirm:Color": 10, "Query:Pattern": 1, 
    "Query:Symbol": 1, "Query:Shape": 1, "Query:Size": 1, "Query:Texture": 1}

    workbook = Workbook()
    sheet = workbook.active

    utilList = []

    # write headers
    col = 1
    for idx, prop in enumerate(propList,start=1):
        sheet.cell(row=1, column=idx).value = prop
        col += 1

    sheet.cell(row=1, column=len(propList)+1).value = "Utility"
    
    # write data
    q_idx = 0
    for q in questions:
        n = 2
        # write cols
        for c in range(1,col):
            # write rows
            if c == 1: # AskQuestion
                for r in range(1,2**(len(propList)-1)+1):
                    sheet.cell(row=r+q_idx+1, column=c).value = q
            elif c == 2: # Refs (/2)
                for r in range(1,2**(len(propList)-1)+1):
                    if r < (2**(len(propList)-1)+1)/n:
                        sheet.cell(row=r+q_idx+1, column=c).value = 2
                    else:
                        sheet.cell(row=r+q_idx+1, column=c).value = 3
                n *= 2
            else: # all others
                count = 1
                maxCount = (2**(len(propList)-1))/n
                known = True
                for r in range(1,2**(len(propList)-1)+1):   
                    # If property relates to question, update Utility column (last col)    
                    # extract "Utility" column into a list     
                    if known: 
                        sheet.cell(row=r+q_idx+1, column=c).value = "Known"
                        if q in propToQ[propList[c-1]]:
                            sheet.cell(row=r+q_idx+1, column=col).value = 0
                            utilList.append(0)
                    else:
                        sheet.cell(row=r+q_idx+1, column=c).value = "Unknown"
                        # use entropy values for each question, e.g., "Query:Color":20.55"
                        if q in propToQ[propList[c-1]]:
                            if entropy:
                                if "Query" in q: # decide which entropy to use (WHQ or YNQ)
                                    ent_val = entropy.WHQentropy[propList[c-1]]
                                    sheet.cell(row=r+q_idx+1, column=col).value = ent_val
                                    utilList.append(ent_val) 
                                else: # "Confirm" in q
                                    ent_val = entropy.YNQentropy[propList[c-1] + "YN"]
                                    sheet.cell(row=r+q_idx+1, column=col).value = ent_val
                                    utilList.append(ent_val) 
                            else: # corpus-based utilities
                                util_val = corpusUtilities[q]
                                sheet.cell(row=r+q_idx+1, column=col).value = util_val
                                utilList.append(util_val) 

                    count += 1
                    # toggle bools
                    if count > maxCount:
                        known = not known
                        count = 1
                n *= 2 #(/4, /8, /16, /32, /64)

        q_idx += 2**(len(propList)-1)

    # save to file
    fname = "UtilityTable.xlsx"
    workbook.save(filename=fname)

    #print(utilList)
    return utilList


def writeBIXMLdistribution(network, var_for, var_given, prior):
    definition = ET.SubElement(network, "DEFINITION")
    f = ET.SubElement(definition, "FOR")
    f.text = var_for
    if var_given is not None:
        for g in var_given:
            given = ET.SubElement(definition, "GIVEN")
            given.text = str(g)
    if prior is not None:
        t = ET.SubElement(definition, "TABLE")  
        t.text = ""  
        for elem in prior:     
            t.text += str(elem) + " "


def writeBIXMLvariable(network, var_type, var_name, var_property, outcomes):
    variable = ET.SubElement(network, "VARIABLE")
    variable.set("TYPE", var_type)
    name = ET.SubElement(variable, "NAME")
    name.text = var_name
    prop = ET.SubElement(variable, "PROPERTY")
    prop.text = var_property
    for o in outcomes:
        outcome = ET.SubElement(variable, "OUTCOME")
        outcome.text = str(o)


# write the bifxml file specifying the decision network using elementTree
def writeBIFXML(minProps,utts,entropy):
    # questions are based on list of questions of each property
    global propToQ
    # create the file structure
    bif = ET.Element('BIF')
    bif.set("VERSION","0.3")
    network = ET.SubElement(bif, "NETWORK")

    allQuestions = [v for k, v in propToQ.items() if k in minProps]
    questions = [item for sublist in allQuestions for item in sublist] # flatten list
    # TODO: sort questions by highest entropy

    # questions (network, TYPE, NAME, PROPERTY, outcomes)
    writeBIXMLvariable(network, "decision", "AskQuestion", "Question", questions)
    # commands
    writeBIXMLvariable(network, "nature", "Command", "Instruction", utts)
    # property nodes: Ask + Refs + ObjType + minimum props
    writeBIXMLvariable(network, "utility", "Ask", "Utility of AskQuestion", [0])
    writeBIXMLvariable(network, "nature", "Refs", "Possible referents", [2,3])
    #writeBIXMLvariable(network, "nature", "ObjectType", "Description", ["Known","Unknown"])
    for prop in minProps:
        writeBIXMLvariable(network, "nature", prop, prop, ["Known","Unknown"])
    

    # write distributions
    # DEFINITION, FOR, GIVEN, TABLE
    prior = [0.5,0.5]
    n = 1/(len(utts))
    commandPrior = []
    for i in range(len(utts)):
        commandPrior.append(str(n))
    writeBIXMLdistribution(network, "Command", None, commandPrior)
    writeBIXMLdistribution(network, "AskQuestion", None, None)
    writeBIXMLdistribution(network, "Refs", None, prior)
    #writeBIXMLdistribution(network, "ObjectType", None, prior)
    for prop in minProps:
        writeBIXMLdistribution(network, prop, None, prior)
    

    # write utility node and table
    givenList = ["AskQuestion","Refs"]
    for prop in minProps:
        givenList.append(prop)
    utilTable = computeUtilityTable(givenList,questions,entropy)
    writeBIXMLdistribution(network, "Ask", givenList, utilTable)

    # create a new XML file with the results
    mydata = ET.tostring(bif)

    xmlstr = minidom.parseString(mydata).toprettyxml(indent="   ")
    num = 1
    with open("decNet" + str(num) + ".bifxml", "w") as myfile:
        myfile.write(xmlstr)

    return myfile

def computeMinimumProperties(world):
    # convert to an instance of the hitting set problem and solve using an approximation algorithm
    # C: [[1, 2, 3], [1, 4], [5, 6, 7]], H: [1, 5]]

    # X: ground set of all properties (shape, size, color, texture, pattern, symbol)
    # C: collection of subsets of properties that each entity differs by e.g., {"Emitter": ["Color", "Size", "Shape"], "Calibrator": [Texture, Size], ...}
    # H: smallest subset that covers C

    # Compute C below:
    H = []
    C = {}
    entities = world.agent.beliefstate.entities
    for e in entities: # TODO: is this outer for loop needed?
        # go through all pairs of ents 
        #entsList = list(combinations(ents, 2)) 
        entsList = list(combinations(entities, 2))
        for ent1, ent2 in entsList:
            C[ent1.getID() + " " + ent2.getID()] = [prop for prop in ent1.properties if ent1.properties[prop] != ent2.properties[prop]]

    # Compute H below
    h = Hitman(solver='m22', htype='rc2')
    for ent, diffList in C.items():
        h.hit(diffList)

    H = h.get()
    print("Smallest set of properties: " + str(H))

    # print all hitting sets
    #with Hitman(bootstrap_with=C.values(), htype='sorted') as hitman:
    #    for hs in hitman.enumerate():
    #        print(hs)

    return H

def constructDecisionNet(world, utts,entropy):
    minProps = computeMinimumProperties(world)
    return writeBIFXML(minProps,utts,entropy)

def processInitialReferent(beliefstate,command,referents):
    # search agent's knowledge about that entity name
    #ref = beliefstate.getEntityByName(command.getReferent())
    print("Referent is: " + str(command.getReferent()))

    # new referent
    if command.getReferent() not in propKnowledge:
        # object or container?
        isObject = False
        if command.getType() == "Object":
            isObject = True 
    
        # prune entities that aren't the same type and that are already known
        for x in beliefstate.entities: # beliefstate.agent.beliefstate.entities
            if isObject and x.properties["EntityType"] == "Object" and x not in referents:
                referents.append(x)
            elif not isObject and x.properties["EntityType"] == "Container" and x not in referents:
                referents.append(x)
        

        # create this entry in the knowledge list (no properties known yet)
        propKnowledge[command.getReferent()] = {}
        print("New object")
        print("propKnowledge is: " + str(propKnowledge))
    
    return len(referents)


# Reference resolution: returns # of refs and updates the referents list of possible entities
def ambiguousReferents(beliefstate,worldstate,command,referents,response):

    print("propKnowledge is: " + str(propKnowledge))
    newReferents = [] # copy so we can prune

    # handle No response to YNQ. 
    if "Confirm" in response.getInput()["QuestionType"] and response.getInput()["Response"] == "No":
        for e in referents: # entity
            if e.properties[response.getInput()["Property"]] != response.getInput()["Value"] and e not in newReferents:
                newReferents.append(e)
    else: # WHQs and Yes response to YNQ      
        # creates list of referents that have at least one of the known properties
        for k, v in propKnowledge.items(): # e.g, "Temporal Emitter":{"ObjectType":"Emitter"}
            for prop, val in v.items(): # dict of Property: Value (e.g., "Color:": "Red")
                for e in referents: # entity
                    if e.properties[prop] == val and e not in newReferents: # e.g., ObjectType == Emitter
                        newReferents.append(e)

    allKnownProps = propKnowledge[command.getReferent()]   

    finalReferents = []
    # prune resulting list so it only contains entities that have ALL properties
    for r in newReferents:
        ref = True
        for k,v in allKnownProps.items(): # "Color: Red"  
            # handle YNQ NO casse
            if "NO" in v:
                split_prop = v[3:] # "NO Red"
                if r.properties[k] == split_prop:
                    ref = False
            # standard case. don't add entity if prop doesn't match
            elif r.properties[k] != v: 
                ref = False
        if ref: 
            finalReferents.append(r)

    # to handle edge cases when refs = 0 because we didn't remove Location matches (happens when agent only knows location)
    if len(finalReferents) > 0:
        print("referents = finalReferents")
        referents = finalReferents

    # TODO: Fix bug: Landmark YNQs will have finalReferents = 0, so this returns the original set of referents    

    print((x for x in referents))
    return len(referents)

def formatQuestion(q, command):
    # mapping of question to property
    qProperty = ""
    if "Location" in str(q): qProperty = "Location"
    elif "Spatial" in str(q): qProperty = "Location"
    elif "Landmark" in str(q): qProperty = "Location"
    elif "ID" in str(q): qProperty = "ObjectType"
    elif "General" in str(q): qProperty = "ObjectType"
    elif "Color" in str(q): qProperty = "Color"
    elif "Size" in str(q): qProperty = "Size"
    elif "Shape" in str(q): qProperty = "Shape"
    elif "Symbol" in str(q): qProperty = "Symbol"
    elif "Pattern" in str(q): qProperty = "Pattern"
    elif "Texture" in str(q): qProperty = "Texture"

    if "Query" in str(q):
        returnQ = luci.WHQ({"QuestionType":str(q), "Referent": command.getReferent(), "Property": qProperty})
    else: #"Confirm" in str(q)
        # randomly choose a value for that property
        qValue = random.choice(luci.objectProperties[qProperty])
        returnQ = luci.YNQ({"QuestionType":str(q), "Referent": command.getReferent(), "Property": qProperty, "Value": qValue}) 

    return returnQ

# call decision net with command, target items, and known props, and get question back
def generateQuestion(command, referents, propKnowledge):
    evs = {"Command": str(command.getInput())}
    if referents > 2: 
        evs["Refs"] = "3"
    else: evs["Refs"] = str(referents)

    for k, v in propKnowledge.items(): # e.g, "Temporal Emitter":{"ObjectType":"Emitter"}
        if k == command.getReferent():
            for prop, val in v.items(): # dict of Property: Value (e.g., "Color:": "Red")
                if "NO" in val and referents > 1:
                    # YNQs No's should be a distribution based on # of refs (e.g., 4 refs; 1 / 4 = [0.25, 0.75])
                    evs[prop] = [1/referents, (referents-1/referents)]
                else:
                    evs[prop] = "Known"

    print("Evidence is: " + str(evs))
    # Call decision net with evidence and return question with MEU
    q = decision_net.bestQuestion(evs)
    
    print("unformatted question is: " + q)

    return formatQuestion(q, command)

def generateQuestionBaseline(command, referents, propKnowledge):
    global propToQ
    # go through propKnowledge and add all UNKNOWN properties to a list
    unknownProps = ["Location", "ObjectType", "Color", "Pattern", "Symbol", "Shape", "Size", "Texture"]
    for k, v in propKnowledge.items(): # e.g, "Temporal Emitter":{"ObjectType":"Emitter"}
        if k == command.getReferent():
            for prop, val in v.items(): # dict of Property: Value (e.g., "Color:": "Red")               
                if not "NO" in val: # we know about it
                    unknownProps.remove(prop)
    
    # randomly select a property from this list
    p = random.choice(unknownProps)

    # find the corresponding question from a dict of prop:q's
    qList = propToQ[p] 

    # format and return that question
    return formatQuestion(random.choice(qList), command)
    

def interpretResponse(command, response):
    global qToProp
    if "Query" in response.getQuestionType():
        if command.getReferent() in propKnowledge:
            # update our knowledge of the referent (this does not update beliefstate because referent is not resolved yet)
            learnedProp = {response.getInput()["Property"]:response.getInput()["Value"]}
            propKnowledge[command.getReferent()].update(learnedProp)
            qToProp += 1
    elif "Confirm" in response.getQuestionType():
        if response.getInput()["Response"] == "Yes":
            # treat it like a WHQ
            learnedProp = {response.getInput()["Property"]:response.getInput()["Value"]}
            propKnowledge[command.getReferent()].update(learnedProp)
            qToProp += 1
        else: # the No case is also handled in ambiguousReferents()
            # only do this if you don't already know it, e.g., if Color in propKnowledge (Crate 2) and something like "Color: NO red" is there
            if [response.getInput()["Property"]] in propKnowledge[command.getReferent()].values() and "NO" in propKnowledge[command.getReferent()][response.getInput()["Property"]]:
                return
            else: # otherwise add the knowledge
                learnedProp = {response.getInput()["Property"]:"NO " + str(response.getInput()["Value"])}
                propKnowledge[command.getReferent()].update(learnedProp) 

            # TODO: propknowledge should really have a list of dicts so you could stack NO info
            # e.g., {Crate 2: {Color: [NO red, NO blue}]
            # it should also be able to learn a property when it has gotten enough No's to narrow it down
            # when there are only a few refs, a No might be enough to learn a prop


# Update knowledge base with the name of the item once reference resolution has completed
def resolveEntity(beliefstate,worldstate,name,referents):
    global propKnowledge, random_world
    # find entity associated with that name from the world state (i.e., "reference resolution")
    if random_world:
        e = worldstate.getEntityByID(name)
    else:
        e = worldstate.getEntityByName(name)
    ID = e.getID()
    beliefstate.getEntityByID(ID).properties["Name"] = name

    # remove entity from referents and belief state
    referents.remove(beliefstate.getEntityByID(ID))
    beliefstate.entities.remove(beliefstate.getEntityByID(ID))

    # if it's a container, update the containedIn field of its content objects (if any) in belief state
    if e.properties["EntityType"] == "Container":
        for i in range(len(e.containedObjects)):
            obj = e.containedObjects[i]
            beliefstate.getEntityByID(obj.getID()).properties["ContainedIn"] = e.getID() # add container to object in belief
            if not beliefstate.getEntityByID(e.getID()).isInside(obj):
                beliefstate.getEntityByID(e.getID()).addObject(obj) # add object to container in belief

    propKnowledge = {} # clear out prop knowledge so it doesn't screw up subsequent input
    
def generateResponse(world,query):
    YNQ = False
    if "Confirm" in query.getInput()["QuestionType"]:
        YNQ = True
        ynqr = "No"
    #print("GenerateResponse")
    ## get relevant entity
    if random_world:
        einfo = world.getEntityByID(query.referent)
    else: 
        einfo = world.getEntityByName(query.referent)
    if query.property in einfo.properties.keys():
        pinfo = einfo.properties[query.property]
        if YNQ and pinfo == query.value:
            ynqr = "Yes"
    else:
        pinfo = "Unknown"
    
    if YNQ:
        return luci.YNQresponse({"QuestionType": query.questionType, "Referent": query.referent, "Property": query.property, "Value": query.value, "Response": ynqr})
    else:
        return luci.WHQresponse({"QuestionType": query.questionType, "Referent": query.referent, "Property": query.property, "Value": pinfo})

# reset question counts
def resetCounts():
    for key in questionCount.keys():
        questionCount[key] = 0

# update question counts
def updateCounts(question):
    q = question.questionType
    if q not in questionCount.keys():
        questionCount[q] = 1
    else:
        questionCount[q] += 1.0

def qa_Run(world,referents,condition,inputUtts):
    global numInputUtts
    results = {}
    numQs = 0
    # Main loop that reads input utterances

    for x in range(len(inputUtts)):
        numInputUtts += 1
        command = inputUtts[x]
        command.debugPrint()
        
        refs = processInitialReferent(world.agent.getBeliefState(),command,referents)

        # possible entities that match unknown referent
        print("Refs: " + str(refs))
 
        while refs != 1: # loop until we narrowed down referents to 1

            if condition == "decision":
                # call decision net to generate question
                q = generateQuestion(command,refs,propKnowledge)
                print("Question from decision net: " + str(q.getInput()))
            else:
                # baseline approach to generate question
                q = generateQuestionBaseline(command,refs,propKnowledge)
                print("Random question: " + str(q.getInput()))

            updateCounts(q)

            # e.g., "Emitters look like med kits" (this would normally come from the commander response)
            response = generateResponse(world,q)
            print("Response: " + str(response.getInput()))

            interpretResponse(command,response)
            # learned the description 

            #print("Learned property")
            refs = ambiguousReferents(world.agent.getBeliefState(),world,command,referents,response)

            print("Refs: " + str(refs))
            numQs += 1

        # Update agent's belief state with name and perform action
        resolveEntity(world.agent.getBeliefState(), world, command.getReferent(),referents)  
        print("Entity resolved" + "\n")

    results["NumberQuestions"] = numQs
    return results


### qa_BatchRun 
# world - ScenarioWorld instance
# condition - QA policy condition; "decision" vs. other scripted policies
# commands - list of input commands
# iterations = number of iterations per command order permutation
def qa_BatchRun(baseWorld,world,condition,commands,iterations):
    global qToProp

    num = 0
    results = {}
    resetCounts()

    for i in range(iterations):
        print ("Iteration: " + str(i))
        qToProp = 0
        #run_results = qa_Run(world,[],condition,list(perm))
        run_results = qa_Run(world,[],condition,commands)

        # clear world knowledge here in between runs
        world.agent.setBeliefState(world.deriveBeliefState(unscannable_props))
        
        ## update partial results
        if "TotalNumberQuestions" not in results.keys():
            results["TotalNumberQuestions"] = []
            results["QsToProp"] = []
        results["TotalNumberQuestions"].append(run_results["NumberQuestions"])
        results["QsToProp"].append(run_results["NumberQuestions"]/qToProp)

        num += 1.0

        # randomize world for next iteration
        if random_world:
            world = initRandomWorld(baseWorld)
            world.agent.setBeliefState(world.deriveBeliefState(unscannable_props))
            initializeDecisionNet(world,commands)

    ## Record Mean Results
    results["MeanQsToProp"] = sum(results["QsToProp"])/len(results["QsToProp"])
    results["STDQsToProp"] = numpy.std(results["QsToProp"])
    results["TotalIterations"] = num
    results["TotalInstructions"] = numInputUtts
    results["InstructsPerIteration"] = numInputUtts/results["TotalIterations"]
    results["QsPerInstruct"] = []
    for qresult in results["TotalNumberQuestions"]:
        results["QsPerInstruct"].append(qresult/results["InstructsPerIteration"])
    results["MeanQsPerInstruct"] = sum(results["TotalNumberQuestions"])/numInputUtts
    results["STDQsPerInstruct"] = numpy.std(results["QsPerInstruct"])

    #results["MeanQsToProp"] = sum(results["TotalNumberQuestions"])/qToProp
    results["MeanQsPerIteration"] = sum(results["TotalNumberQuestions"])/results["TotalIterations"]
    results["STDQsPerIteration"] = numpy.std(results["TotalNumberQuestions"])
    return results


def parseInputUttsRandom():
    inputUtts = [ luci.Command({'Referent': 'obj0', 'Type': 'Object'}), 
    luci.Command({'Referent': 'obj1', 'Type': 'Object'}), 
    luci.Command({'Referent': 'obj2', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj3', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj4', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj5', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj6', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj7', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj8', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj9', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj10', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj11', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj12', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj13', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj14', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj15', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj16', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj17', 'Type': 'Object'}), 
    luci.Command({'Referent': 'obj18', 'Type': 'Object'}),
    luci.Command({'Referent': 'obj19', 'Type': 'Object'}) ]

    return inputUtts

# TODO: Parse from input file and store as dict of Commands
def parseInputUtts():
    inputUtts = [ luci.Command({'Referent': 'Cryo calibrator', 'Type': 'Object'}), 
    luci.Command({'Referent': 'Delta calibrator', 'Type': 'Object'}), 
    luci.Command({'Referent': 'Mechano calibrator', 'Type': 'Object'}),
    luci.Command({'Referent': 'Temporal emitter', 'Type': 'Object'}),
    luci.Command({'Referent': 'Organic emitter', 'Type': 'Object'}),
    luci.Command({'Referent': 'Plasma emitter', 'Type': 'Object'}),
    luci.Command({'Referent': 'Electro capacitor', 'Type': 'Object'}),
    luci.Command({'Referent': 'Adaptive capacitor', 'Type': 'Object'}),
    luci.Command({'Referent': 'Tesla capacitor', 'Type': 'Object'}),
    luci.Command({'Referent': 'Megaband module', 'Type': 'Object'}),
    luci.Command({'Referent': 'Hyperband module', 'Type': 'Object'}),
    luci.Command({'Referent': 'Ultraband module', 'Type': 'Object'}),
    luci.Command({'Referent': 'Modular synthesizer', 'Type': 'Object'}),
    luci.Command({'Referent': 'Optimized synthesizer', 'Type': 'Object'}),
    luci.Command({'Referent': 'Galvanic synthesizer', 'Type': 'Object'}),
    luci.Command({'Referent': 'Sonic optimizer', 'Type': 'Object'}),
    luci.Command({'Referent': 'Advanced optimizer', 'Type': 'Object'}),
    luci.Command({'Referent': 'Quantum optimizer', 'Type': 'Object'}),
    luci.Command({'Referent': 'Crate 1', 'Type': 'Object'}),
    luci.Command({'Referent': 'Crate 2', 'Type': 'Object'}) ]
        
    return inputUtts

def initializeDecisionNet(world,inputUtts):
    if condition == "decision":
        if utilities == "entropy":
            WHQentropy, YNQentropy = UtilityCalculation.main(world)
            entropy = Entropy(WHQentropy, YNQentropy)
            filename = constructDecisionNet(world,inputUtts,entropy)
        else: # corpus
            entropy = None
            filename = constructDecisionNet(world,inputUtts,entropy)
        decision_net.initializeNetwork(filename.name)

def initRandomWorld(baseWorld):
    randomWorld = luci.ScenarioWorld()
    objects = randomize_world.main(baseWorld,complexity)
    for obj in objects:
        randomWorld.addRandomEntity(obj)
        
    randomWorld.initAgent()
    randomWorld.agent.setBeliefState(randomWorld.deriveBeliefState(unscannable_props))

    return randomWorld

def main():
    global random_world, complexity, condition, utilities

    # list of Entities
    referents = []

    # list of commands
    inputUtts = parseInputUtts()
    #sampleUtts = random.sample(inputUtts, 2) # use subset of utts from list

    # input world
    world = luci.ScenarioWorld()
    if random_world:
        world.parseFromCSV("<path_to>/evaluation/TestWorld_eval_uniform.csv")
    else:
        world.parseFromCSV("<path_to>/evaluation/TestWorld_eval.csv")
    world.agent.setBeliefState(world.deriveBeliefState(unscannable_props))
    baseWorld = world
     

    # randomize world and get objects
    if random_world:
        world = initRandomWorld(world)
        inputUtts = parseInputUttsRandom()
        world.agent.setBeliefState(world.deriveBeliefState(unscannable_props))

    initializeDecisionNet(world,inputUtts)

    results = qa_BatchRun(baseWorld,world,condition,inputUtts,100)

    print("Random world: " + str(random_world) + ", Complexity: " + complexity + ", Condition: " + condition + ", Utilities: " + utilities)
    print(questionCount)
    print(results)
    

if __name__ == "__main__":
    main()
